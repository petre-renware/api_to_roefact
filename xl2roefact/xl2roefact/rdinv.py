#!../.venv/bin/python3
"""rdinv: modul de procesare a fisierului Excel ce contine factura si colectare a datelor aferente.

Formatul acceptat fisier Excel este `XLSX`.

Identification:

* code-name: `rdinv`
* copyright: (c) 2023 RENWare Software Systems
* author: Petre Iordanescu (petre.iordanescu@gmail.com)

Specifications:

* document cerinte initiale: `110-SRE-api_to_roefact_requirements.md` section `Componenta xl2roefact`
* INTRARI: fisier format XLSX ce contine factura emisa (cod: `f-XLSX`)
* IESIRI: fisier format JSON imagine a datelor facturii (cod: `f-JSON`)
"""

import os, sys
from datetime import datetime, timezone, tzinfo
from rich import print
import copy
from rich.pretty import pprint
from string import ascii_lowercase
import json
from typing import Callable, Any
from functools import partial
import pylightxl as xl
import openpyxl as opnxl

# local modules (part of package), application misc/general utilities
from .libutils import isnumber
from .libutils import find_str_in_list
from .libutils import dict_sum_by_key
from . import config_settings  # application configuration parameters

# local constants. Change them with caution only for a functional objective
SYS_FILLED_EMPTY_CELL = "_sys_keep_cell"

# imported constants (NOTE: some are subject to change values as reading Excel file - these will be declared as `global` in function that will change them)
DEFAULT_VAT_PERCENT = config_settings.DEFAULT_VAT_PERCENT
DEFAULT_UNKNOWN_ITEM_NAME = config_settings.DEFAULT_UNKNOWN_ITEM_NAME
DEFAULT_UNKNOWN_UOM = config_settings.DEFAULT_UNKNOWN_UOM
DEFAULT_CURRENCY = config_settings.DEFAULT_CURRENCY
DEFAULT_CUSTOMER_COUNTRY = config_settings.DEFAULT_CUSTOMER_COUNTRY
DEFAULT_SUPPLIER_COUNTRY = config_settings.DEFAULT_SUPPLIER_COUNTRY
PATTERN_FOR_INVOICE_ITEMS_SUBTABLE_MARKER = config_settings.PATTERN_FOR_INVOICE_ITEMS_SUBTABLE_MARKER
PATTERN_FOR_INVOICE_NUMBER_LABEL = config_settings.PATTERN_FOR_INVOICE_NUMBER_LABEL
PATTERN_FOR_INVOICE_CURRENCY_LABEL = config_settings.PATTERN_FOR_INVOICE_CURRENCY_LABEL
PATTERN_FOR_INVOICE_ISSUE_DATE_LABEL = config_settings.PATTERN_FOR_INVOICE_ISSUE_DATE_LABEL
PATTERN_FOR_INVOICE_SUPPLIER_SUBTABLE_MARKER = config_settings.PATTERN_FOR_INVOICE_SUPPLIER_SUBTABLE_MARKER
PATTERN_FOR_INVOICE_CUSTOMER_SUBTABLE_MARKER = config_settings.PATTERN_FOR_INVOICE_CUSTOMER_SUBTABLE_MARKER
PATTERN_FOR_PARTNER_ID = config_settings.PATTERN_FOR_PARTNER_ID
PATTERN_FOR_CUSTOMER_LEGAL_NAME = config_settings.PATTERN_FOR_CUSTOMER_LEGAL_NAME
PATTERN_FOR_PARTNER_ADDRESS = config_settings.PATTERN_FOR_PARTNER_ADDRESS
PATTERN_FOR_PARTNER_ADDRESS_COUNTRY = config_settings.PATTERN_FOR_PARTNER_ADDRESS_COUNTRY
PATTERN_FOR_PARTNER_ADDRESS_CITY = config_settings.PATTERN_FOR_PARTNER_ADDRESS_CITY
PATTERN_FOR_PARTNER_ADDRESS_STREET = config_settings.PATTERN_FOR_PARTNER_ADDRESS_STREET
PATTERN_FOR_PARTNER_ADDRESS_ZIPCODE = config_settings.PATTERN_FOR_PARTNER_ADDRESS_ZIPCODE
PATTERN_FOR_PARTNER_EMAIL = config_settings.PATTERN_FOR_PARTNER_EMAIL
PATTERN_FOR_PARTNER_TEL = config_settings.PATTERN_FOR_PARTNER_TEL
PATTERN_FOR_PARTNER_IBAN = config_settings.PATTERN_FOR_PARTNER_IBAN
PATTERN_FOR_PARTNER_BANK = config_settings.PATTERN_FOR_PARTNER_BANK
PATTERN_FOR_PARTNER_REGCOM = config_settings.PATTERN_FOR_PARTNER_REGCOM


def rdinv(
        file_to_process: str,
        invoice_worksheet_name: str = None,
        *,
        debug_info: bool = False
    ) -> dict:
    """read Excel file for invoice data.

    Produce a dictionary structure + JSON file with all data regarding read invoice: canonical KV data, meta data, map to convert to XML and original Excel data.

    Args:
        `file_to_process`: the invoice file (exact file with path).
        `invoice_worksheet_name`: the worksheet containing invoice, optional, defaults to first found worksheet.
        `debug_info`: key only, show debugging information, default `False`.

    Return:
        `dict`: the invoice extracted information from Excel file as `dict(Invoice: dict, meta_info: dict, excel_original_data: dict)`  #TODO subject of documentation update.

    NOTE ref important variables:
        * `db: pylightxl object`: EXCEL object with invoice (as a whole)
        * `ws: pylightxl object`: WORKSHEET object with invoice
    """
    # use as global only for those constants that could be changed by this function
    global DEFAULT_VAT_PERCENT
    global DEFAULT_UNKNOWN_ITEM_NAME
    global DEFAULT_UNKNOWN_UOM
    global DEFAULT_CURRENCY
    global DEFAULT_SUPPLIER_COUNTRY
    global DEFAULT_CUSTOMER_COUNTRY

    print(f"*** Module [red]rdinv[/] started at {datetime.now()} to process file [green]{os.path.split(file_to_process)[1]}[/] (full path: {file_to_process})")

    # read Excel file with Invoice data
    try:
        db = xl.readxl(fn=file_to_process)
    except:
        print(f"[red]---***FATAL ERROR - Cannot open Excel file {file_to_process} (possible problems: file corrupted, wrong type only XLSX accetpted, file does not exists or was deleted, operating system vilotation) in Module [red] RDINV (code-name: `rdinv`). File processing terminated[/]")
        return False

    # read the workshet with Invoice data
    if invoice_worksheet_name is None:  # if parameter `invoice_worksheet_name` not specified try to open first worksheet from Excel worksheets - order is given by worksheets order in Excel file
        list_of_excel_worksheets = db.ws_names
        print(f"[yellow]INFO note:[/] `rdinv` module, no worksheet specified, will open first one: [cyan]'{list_of_excel_worksheets[0]}'[/]")
        invoice_worksheet_name = list_of_excel_worksheets[0]

    try:
        ws = db.ws(invoice_worksheet_name)
    except:
        print(f"[red]***FATAL ERROR - Cannot open Excel specified Worksheet \"{invoice_worksheet_name}\" in Module [red] RDINV (code-name: `rdinv`). File processing terminated[/]")
        return False

    """#NOTE: section for search of `invoice_items_area: pylightxl.ssd`
        - main result: `keyword_for_items_table_marker` = string marker to search for in oredr to isolate `invoice_items_area`
        - other results: `_found_cell_for_invoice_items_area_marker = (row, col, val)`
    """
    _tmp_label_info = get_excel_data_at_label(
        pattern_to_search_for=PATTERN_FOR_INVOICE_ITEMS_SUBTABLE_MARKER,
        worksheet=ws,
        targeted_type=str
    )
    if _tmp_label_info["label_value"]:  # only if found a potential cell usable as`invoice_items_area_marker`
        _found_cell_for_invoice_items_area_marker = (
            _tmp_label_info["label_location"][0],
            _tmp_label_info["label_location"][1],
            _tmp_label_info["label_value"]
        )  # here you still need to construct `_found_cell_for_invoice_items_area_marker = (row, col, val)` as it is used in prev code (before `231220piu_a` change)
    else:
        print(f"[red]***FATAL ERROR - Cannot find a relevant cell where invoice items table start (basically containing string \" crt\"). File processing terminated[/]")
        return False
    keyword_for_items_table_marker = _found_cell_for_invoice_items_area_marker[2]

    # detect all cells that should be filled with SYS_FILLED_EMPTY_CELL (these are cells id merged groups where first cell in merged group is relevant (diff from empty))
    detected_cells_which_will_be_fake_filled = _get_merged_cells_tobe_changed(
        file_to_scan=file_to_process,
        invoice_worksheet_name=invoice_worksheet_name,
        keep_cells_of_items_ssd_marker=_found_cell_for_invoice_items_area_marker)  # this call specify to keep unchanged that cells with some description
    for _cell_index in detected_cells_which_will_be_fake_filled:  # scan all detectected cell and change them
        _cell_row = _cell_index[0]
        _cell_col = _cell_index[1]
        ws.update_index(row = _cell_row, col = _cell_col, val = SYS_FILLED_EMPTY_CELL)

    """#NOTE: section to "solve" `invoice_items_area`. Steps:
        - process it as Excel format (row & colymns tabular organization)
        - transform it in "canonical JSON format" (as kv pairs) and update `cac_InvoiceLine` key
    """
    # process invoice to detect its items / lines ('invoice_items_area'), clean and extract data
    invoice_items_area = get_invoice_items_area(
        worksheet=ws,
        invoice_items_area_marker=keyword_for_items_table_marker,
        wks_name=invoice_worksheet_name
    )

    """#NOTE: section for localize areas: invoice header (`invoice_header_area`) & invoice footer (`invoice_footer_area`)
        NOTE: its code suppose `invoice_items_area` is already defined (as location). It will be used as reference in "header" & "footer" localization logic
    """
    ulc_header = (1, 1)
    lrc_header = (_found_cell_for_invoice_items_area_marker[0] - 1, ws.size[1],)  # info where header ends: row = from items data table start location `_found_cell_for_invoice_items_area_marker[0]-1` && col = las col of worksheet
    invoice_header_area = dict(
        start_cell = ulc_header,
        end_cell = lrc_header
    )
    #
    _ulc_footer_row = _found_cell_for_invoice_items_area_marker[0] + len(invoice_items_area["keyrows"]) + 1  # this is located at: row = items data table start row  (_found_cell_for_invoice_items_area_marker[0]) + # of items data table rows + 1 (invoice_items_area["data"]) && col = 1
    _start_cell_val = ws.index(_ulc_footer_row, 1)
    while (_start_cell_val != "") and (_ulc_footer_row <= ws.size[0]): # but if that cell is not blank repeatedly to go une more row down... (why could happen? because ietms data table header was composed of more rows by merging cells...)
        _ulc_footer_row += 1
        _start_cell_val = ws.index(_ulc_footer_row, 1)
    ulc_footer = (_ulc_footer_row, 1)
    lrc_footer =( ws.size[0], ws.size[1])  # end of worksheet
    invoice_footer_area = dict(
        start_cell = ulc_footer,
        end_cell = lrc_footer
    )

    """#NOTE: section to "solve" `invoice_header_area`.
            The kind of info expected in this area: invoice number,  currency, issued date, supplier data, customer data)
    """
    invoice_header_area = invoice_header_area | dict(  # build effective data area & merge localization info from initial dict creation
        invoice_number = None,
        issued_date = None,
        currency = None,
        customer_area = None,
        supplier_area = "...future..."  # TODO: ... future tbd  ...
    )  #FIXME_TODO: ............hereuare............
    _area_to_search = (invoice_header_area["start_cell"], invoice_header_area["end_cell"])  # this is "global" for this section (corners of `invoice_header_area`)
    #
    # find invoice number ==> `cbc:ID`
    invoice_number_info = get_excel_data_at_label(
        pattern_to_search_for=PATTERN_FOR_INVOICE_NUMBER_LABEL,
        worksheet=ws,
        area_to_scan=_area_to_search,
        targeted_type=str
    )  # returned info: `{"value": ..., "location": (row..., col...)}`
    invoice_header_area["invoice_number"] = copy.deepcopy(invoice_number_info)
    #
    # find invoice currency ==> `cbc:DocumentCurrencyCode`
    invoice_currency_info = get_excel_data_at_label(
        pattern_to_search_for=PATTERN_FOR_INVOICE_CURRENCY_LABEL,
        worksheet=ws,
        area_to_scan=_area_to_search,
        targeted_type=str
    )  # returned info: `{"value": ..., "location": (row..., col...)}`
    if (invoice_currency_info["value"] is not None) and (invoice_currency_info["value"] != ""):  # if found a currency MUST CHANGE `DEFAULT_CURRENCY` to be properly used for `invoice_items_area`
        DEFAULT_CURRENCY = invoice_currency_info["value"]
    invoice_header_area["currency"] = copy.deepcopy(invoice_currency_info)
    #
    # find invoice issued date ==> `cbc:IssueDate`
    issued_date_info = get_excel_data_at_label(
        pattern_to_search_for=PATTERN_FOR_INVOICE_ISSUE_DATE_LABEL,
        worksheet=ws,
        area_to_scan=_area_to_search,
        targeted_type=str
    )  # returned info: `{"value": ..., "location": (row..., col...)}`
    issued_date_info["value"] = issued_date_info["value"].replace("/", "-")  # convert from Excel format: YYYY/MM/DD (ex: 2023/08/28) to required format in XML file is: `YYYY-MM-DD` (ex: 2013-11-17)
    invoice_header_area["issued_date"] = copy.deepcopy(issued_date_info)
    #
    # find invoice customer ==> "cac:AccountingCustomerParty
    invoice_customer_info = get_excel_data_at_label(
        pattern_to_search_for=PATTERN_FOR_INVOICE_CUSTOMER_SUBTABLE_MARKER,
        worksheet=ws,
        area_to_scan=(invoice_header_area["start_cell"], invoice_header_area["end_cell"]),
        targeted_type=str
    )  # returned info: `{"value": ..., "location": (row..., col...)}`
    # set a dedicated AREAs TO SEARCH for customer
    _area_to_search_start_cell = [  # use `label_location` as being supposed "most far away" from effective-good info, so more chances to find info
        0 if invoice_customer_info["label_location"][0] <= 0 else invoice_customer_info["label_location"][0] - 1,  # set one line up if this line exists
        invoice_customer_info["label_location"][1],
    ]
    if ws.index(*_area_to_search_start_cell).strip() == "":  # prev set was for one line up but if that cell is blank remake it (ie, do a +1)
        _area_to_search_start_cell[0] += 1
    # from `_area_to_search_start_cell` go down up a blank (empty cell)
    _last_ok_position = list([0, 0])
    for __i in range(_area_to_search_start_cell[0], ws.size[0] + 1):  # scan rest of lines for a blank one
        _crt_scanned_cell_idx = (__i, _area_to_search_start_cell[1])
        _crt_scanned_cell_val = ws.index(*_crt_scanned_cell_idx)
        if _crt_scanned_cell_val.strip() == "":  # here you must stop
            break
        # save current position to be used after break
        _last_ok_position = copy.deepcopy(_crt_scanned_cell_idx)
    _area_to_search_end_cell = [
        _last_ok_position[0],
        ws.size[1] if _last_ok_position[1] > ws.size[1] else _last_ok_position[1] + 1,  # set one row right if this row exists
    ]
    # set-persist `_area_to_search` for next steps & save its key-info in associated invoice JSON (for further references)
    _area_to_search = (tuple(_area_to_search_start_cell), tuple(_area_to_search_end_cell))
    invoice_header_area["customer_area"] = {
        "area_info": {
            "value": ws.index(*_area_to_search[0]),  # ie, the value at area start position
            "location": copy.deepcopy(_area_to_search),
        }
    }
    #
    # find customer key "CUI / Registration ID" ==> `invoice_header_area...[CUI]` && `Invoice...[cbc_CompanyID]`
    _temp_found_data = get_excel_data_at_label(
        pattern_to_search_for=PATTERN_FOR_PARTNER_ID,
        worksheet=ws,
        area_to_scan=_area_to_search,
        targeted_type=str,
        down_search_try=False  # customer area is supposed to be organized as "label & value @ RIGHT" or "label: value @ IN-LABEL" but never @ DOWN as being a "not-a-practiced-natural-way"
    )  # returned info: `{"value": ..., "location": (row..., col...)}`
    invoice_header_area["customer_area"]["CUI"] = {
        "value": _temp_found_data["value"],
        "location": _temp_found_data["location"],
        "label_value": _temp_found_data["label_value"],
        "label_location": _temp_found_data["label_location"]
    }
    #
    # find customer key "RegistrationName" ==> `cbc_RegistrationName`
    '''#NOTE: `ReNaSt`-RegNameStrategy (remark: step codes will referred as defined here)
          ReNaSt.STEP-1. search for PATTERN_FOR_CUSTOMER_LEGAL_NAME
          ReNaSt.STEP-2. if `label_location` of FOUND VALUE has the same location as `invoice_header_area["customer_area"]["area_info"]["location"][0]`:
                             keep VALUE of FOUND info
          ReNaSt.STEP-3. else:
                             keep `invoice_header_area["customer_area"]["area_info"]["value"]`
    '''
    _temp_found_data = get_excel_data_at_label(  # NOTE: ReNaSt.STEP-1
        pattern_to_search_for=PATTERN_FOR_CUSTOMER_LEGAL_NAME,
        worksheet=ws,
        area_to_scan=_area_to_search,
        targeted_type=str,
        down_search_try=True  # NOTE: set on True to obtain identical results as original search of `PATTERN_FOR_INVOICE_CUSTOMER_SUBTABLE_MARKER` because name is supposed to be in a very "unstructured mode"
    )  # returned info: `{"value": ..., "location": (row..., col...)}`
    _location_of_header_partner_area = invoice_header_area["customer_area"]["area_info"]["location"][0]
    _location_of_value_found = _temp_found_data["label_location"]
    if _location_of_value_found == _location_of_header_partner_area:  # NOTE: ReNaSt.STEP-2
        kept_RegistrationName = _temp_found_data["value"]
        kept_RegistrationName_location = _temp_found_data["location"]
    else:  # NOTE: ReNaSt.STEP-3
        kept_RegistrationName = invoice_header_area["customer_area"]["area_info"]["value"]
        kept_RegistrationName_location = invoice_header_area["customer_area"]["area_info"]["location"][0]
    invoice_header_area["customer_area"]["RegistrationName"] = {
        "value": kept_RegistrationName,
        "location": kept_RegistrationName_location,
        "label_value": "n/a",
        "label_location": "n/a"
    }
    #
    # find customer key `cac:PostalAddress` -> `invoice_header_area["cac_PostalAddress"]` && Invoice...["cac_PostalAddress"]
    _temp_found_data = get_excel_data_at_label(
        pattern_to_search_for=PATTERN_FOR_PARTNER_ADDRESS,
        worksheet=ws,
        area_to_scan=_area_to_search,
        targeted_type=str,
        down_search_try=False  # customer area is supposed to be organized as "label & value @ RIGHT" or "label: value @ IN-LABEL" but never @ DOWN as being a "not-a-practiced-natural-way"
    )  # returned info: `{"value": ..., "location": (row..., col...)}`
    _tmpstr = _temp_found_data["label_value"].lower()
    _val_is_full_addr = ("adr" in _tmpstr) or ("addr" in _tmpstr)
    if _val_is_full_addr:
        area_to_scan_address_items = (_temp_found_data["location"], _temp_found_data["location"])
    else:
        area_to_scan_address_items = _area_to_search
    search_address_parts = partial(  # define a partial function to be used for all address items search
        get_excel_data_at_label,  # function to call
        worksheet=ws,
        area_to_scan=area_to_scan_address_items,
        targeted_type=str,
        down_search_try=False  # customer area is supposed to be organized as "label & value @ RIGHT" or "label: value @ IN-LABEL" but never @ DOWN as being a "not-a-practiced-natural-way"
    )  # returned info: `{"value": ..., "location": (row..., col...)}`
    _tmp_country = str(search_address_parts(pattern_to_search_for=PATTERN_FOR_PARTNER_ADDRESS_COUNTRY)["value"]).replace("None", "").strip()
    _tmp_city = str(search_address_parts(pattern_to_search_for=PATTERN_FOR_PARTNER_ADDRESS_CITY)["value"]).replace("None", "").strip()
    _tmp_street = str(search_address_parts(pattern_to_search_for=PATTERN_FOR_PARTNER_ADDRESS_STREET)["value"]).replace("None", "").strip()
    _tmp_zipcode = str(search_address_parts(pattern_to_search_for=PATTERN_FOR_PARTNER_ADDRESS_ZIPCODE)["value"]).replace("None", "").strip()
    if (_tmp_country is None) or (_tmp_country == ""):
        _tmp_country = DEFAULT_CUSTOMER_COUNTRY
    else:
        DEFAULT_CUSTOMER_COUNTRY = _tmp_country  # update deflaute value to be re-used in other parts if neccesary
    invoice_header_area["customer_area"]["PostalAddress"] = {
        "cbc_StreetName": _tmp_street,
        "cbc_CityName": _tmp_city,
        "cbc_PostalZone": _tmp_zipcode,
        "cac_Country": {"cbc_IdentificationCode": _tmp_country},
    }
    #
    # search_extended_parts: rest of keys, like: "reg com", "bank / IBAN / cont", "tel", "email" (in code will use names like this: "search_extended_parts")*
    search_extended_parts = partial(  # define a partial function to be used for all "search_extended_parts"
        get_excel_data_at_label,  # function to call
        worksheet=ws,
        area_to_scan=_area_to_search,  # supposed to still contain customer info found area
        targeted_type=str,
        down_search_try=False  # customer area is supposed to be organized as "label & value @ RIGHT" or "label: value @ IN-LABEL" but never @ DOWN as being a "not-a-practiced-natural-way"
    )  # returned info: `{"value": ..., "location": (row..., col...)}`
    _tmp_reg_com = search_extended_parts(pattern_to_search_for=PATTERN_FOR_PARTNER_REGCOM)
    _tmp_bank = search_extended_parts(pattern_to_search_for=PATTERN_FOR_PARTNER_BANK)
    _tmp_IBAN = search_extended_parts(pattern_to_search_for=PATTERN_FOR_PARTNER_IBAN)
    _tmp_phone = search_extended_parts(pattern_to_search_for=PATTERN_FOR_PARTNER_TEL)
    _tmp_email = search_extended_parts(pattern_to_search_for=PATTERN_FOR_PARTNER_EMAIL)
    # store "full" variables in `customer_area...` for excel original values
    invoice_header_area["customer_area"]["reg_com"] = _tmp_reg_com
    invoice_header_area["customer_area"]["bank"] = _tmp_bank
    invoice_header_area["customer_area"]["IBAN"] = _tmp_IBAN
    invoice_header_area["customer_area"]["phone"] = _tmp_phone
    invoice_header_area["customer_area"]["email"] = _tmp_email


    # NOTE: see how replicate code for Customer --to--> Supplier
    # NOTE: mai sunt ai cele "pre-stabilite" in versiunea curenta, gen `cbc:InvoiceTypeCode = 380`
    # NOTE: si mai este ceva legat de o sumarizare XML a totalului facturi (comentarii in zona in care scrii key Invoice, citeva linii mai jos)
    ''' #FIXME ----------------- END OF section for solve `invoice_header_area` (started on line 158) '''


    """#NOTE: section to ( Excel data )--->( JSON ) format preparation and finishing
        this is required to be after header determination (because CURRENCY could be known here and will impact config param `DEFAULT_CURRENCY`)
    """
    # transform `invoice_items_area` in "canonical JSON kv pairs format" (NOTE this step is done only for invoice_items_area and is required because this section is "table with more rows", ie, not a simple key-val)
    invoice_items_as_kv_pairs = mk_kv_invoice_items_area(invoice_items_area_xl_format=invoice_items_area)

    # preserve processed Excel file meta information: start address, size.
    meta_info = _build_meta_info_key(
        excel_file_to_process=file_to_process,
        invoice_worksheet_name=invoice_worksheet_name,
        ws_size=tuple(ws.size),
        keyword_for_items_table_marker=keyword_for_items_table_marker,
        found_cell=tuple(_found_cell_for_invoice_items_area_marker))

    # build final structure to be returned (`invoice`) - MAIN OBJECTIVE of this function
    tmp_InvoiceLine_list = [_i for _i in invoice_items_as_kv_pairs][0],  # first item is `invoice_items_as_kv_pairs` is list of dicts with keys as XML RO E-Fact standard
    print(f"\n=================== tmp_InvoiceLine_list is \n{tmp_InvoiceLine_list}\n")  #FIXME DBG drop me
    invoice = {
        "Invoice": {
            "cbc_ID": copy.deepcopy(invoice_header_area["invoice_number"]["value"]),  # invoice number as `cbc_ID`
            "cbc_DocumentCurrencyCode": copy.deepcopy(invoice_header_area["currency"]["value"]),  # invoice currency as `cbc_DocumentCurrencyCode`
            "cbc_IssueDate": copy.deepcopy(invoice_header_area["issued_date"]["value"]),  # invoice issue date as `cbc_IssueDate`
            "cac_AccountingCustomerParty": {
                "cac:Party": {
                    "cac_PartyLegalEntity": {
                        "cbc_CompanyID": copy.deepcopy(invoice_header_area["customer_area"]["CUI"]["value"]),
                        "cbc_RegistrationName": copy.deepcopy(invoice_header_area["customer_area"]["RegistrationName"]["value"]),
                    },
                    "cac_PostalAddress": copy.deepcopy(invoice_header_area["customer_area"]["PostalAddress"]),
                    "cac_Contact": {
                        "cbc_Telephone": copy.deepcopy(invoice_header_area["customer_area"]["phone"]["value"]),
                        "cbc_ElectronicMail": copy.deepcopy(invoice_header_area["customer_area"]["email"]["value"]),
                        "RegCom": copy.deepcopy(invoice_header_area["customer_area"]["reg_com"]["value"]),
                        "Bank": copy.deepcopy(invoice_header_area["customer_area"]["bank"]["value"]),
                        "IBAN": copy.deepcopy(invoice_header_area["customer_area"]["IBAN"]["value"]),
                    },
                },
            },
            "cac_InvoiceLine": copy.deepcopy(tmp_InvoiceLine_list),  #FIXME @240223 04:00 is list[listtdict]]. Expected list[dict]. Symptom: The effective list is included in other useless list.


            #FIXME: ...hereuare... after finish `invoice_header_area` need  to contsruct TOTAL invoice structure (see #NOTE: "TOTAL_invoice_strucuture")
            "cac_LegalMonetaryTotal": {
                "cbc_LineExtensionAmount": "...",  # ROUND...SUM  dict_sum_by_key(tmp_InvoiceLine_dict, "cbc_LineExtensionAmount")
                "cbc_TaxExclusiveAmount": "...",  #  ROUND...SUM   (`cac_InvoiceLine.cbc_LineExtensionAmount`)
                "cbc_TaxInclusiveAmount": "...",  #  ROUND...SUM   (`cac_InvoiceLine.cbc_LineExtensionAmount` + `cac_InvoiceLine.LineVatAmount`)
                "cbc_PayableAmount": "...",  #       ROUND...SUM   (`cac_InvoiceLine.cbc_LineExtensionAmount` + `cac_InvoiceLine.LineVatAmount`)
            },
            #FIXME ...END of ...hereuare...



        },
        "meta_info": copy.deepcopy(meta_info),
        "excel_original_data": dict(
            invoice_items_area = copy.deepcopy(invoice_items_area),  # NOTE ready, test PASS @ 231205 by [piu]
            invoice_header_area = copy.deepcopy(invoice_header_area),  #TODO wip...
            invoice_footer_area = copy.deepcopy(invoice_footer_area)  #TODO to be done... (just localized `invoice_footer_area`)
        )
    }
    #
    # write `invoice` dict to `f-JSON`
    """ useful NOTE(s):
        - ref `f-JSON` file, see doc: `https://apitoroefact.renware.eu/commercial_agreement/110-SRE-api_to_roefact_requirements.html#vedere-de-ansamblu-a-solutiei`
        - create `f-JSON` filename as Excel filename but with json extention
        - helpers:
                - os.path.split() gets @[0] directory & @[1] filename
                - os.path.splitext() @[0] filename w/o ext, @[1] extention woth dot char included
        - TODO: @231217 - writing a JSON file should be subject of option bool parameter in rdinv() with default value `True`
    """
    _fjson_filename = os.path.splitext(os.path.basename(file_to_process))[0] + ".json"
    _fjson_fileobject = os.path.join(os.path.split(file_to_process)[0], _fjson_filename)
    with open(_fjson_fileobject, 'w', encoding='utf-8') as _f:
        json.dump(invoice, _f, ensure_ascii = False, indent = 4)
    print(f"[yellow]INFO note:[/] `rdinv` module, written invoice JSON data to: [green]{_fjson_fileobject}[/]")

    #TODO check for more TODOs, clean &&-->
    #TODO wip...(@231125) TRANSFORM JSON FILE from Excel (row,col) format in a relational one (but respecting ROefact tags from used scheme)

    return copy.deepcopy(invoice)






# #NOTE - ready, test PASS @ 231212 by [piu]
def get_excel_data_at_label(
        pattern_to_search_for: list[str],
        worksheet: xl.Database.ws,
        area_to_scan: list[list[int]] = None,
        targeted_type: Callable = str,
        down_search_try: bool = True
    ) -> dict:
    """get "one key Excel values", like invoice number or invoice issue date.

    Args:
        `pattern_to_search_for`: for example for inv number, will pass the `PATTERN_FOR_INVOICE_NUMBER_LABEL`.
        `worksheet`: the worksheet containing invoice (as object of `pyxllight` library).
        `area_to_scan`: area of cells to be searched, default whole worksheet.
        `targeted_type`: what type expect (will try to convert to, if cannot will return str), default `str`.
        `down_search_try`: establish if DOWN search method is tried, default `True`.

    Return:
        `None` if not found OR `dictionary` containing:
            * `"value": int | float | str` - the value found covenrted to requested `targeted_type` if possible or `str` otherwise; if "out of space" then returns `None`
            * `"location": (row, col)` - adrees of cell where found value

    Notes:
        * normal scan order is 1.RIGHT, 2.DOWN (if allowed), 3.IN-LABEL only in given area and pattern.
    """
    def __check_value(val: Any) -> bool:
        """ return `True` if a `val` is different of None or empty string or SYS_FILLED_EMPTY_CELL, otherwise return `False`
        """
        if val is None:
            return False
        if isinstance(val, str) and val.strip() == "":
            return False
        if val == SYS_FILLED_EMPTY_CELL:
            return False
        return True

    if area_to_scan is None:  # if arg "passed" was default value then make it as the whole worksheet area
        area_to_scan = ((1, 1), (worksheet.size[0], worksheet.size[1]))
    ret_val = dict(  # initialize structure for what will return if information is found
        value = None,
        location = (None, None),
        label_value = None,
        label_location = None,
    )
    for i in range(area_to_scan[0][0], area_to_scan[1][0] + 1):
        for j in range(area_to_scan[0][1], area_to_scan[1][1] + 1):
            _crt_cell_val = [worksheet.index(i, j)]  # make it list to be iterable (of 1 element, but iterable)
            # test if crt cell is in pattern_to_search_for
            _found = find_str_in_list(list_of_str_to_find=pattern_to_search_for, list_to_search=_crt_cell_val)
            if _found is not None:
                if True:  # process label
                    label_value = _crt_cell_val[0]  # preserve only first value (this will be next proccessed and if pass will remain)
                    label_location = (i, j)
                if True:  # process value
                    value_found = None
                    index_of_value_found = None
                    # NOTE-LOGIC: test for RIGHT cell @(i,j+1) if cell exists in ws range AND if has a value then continue loop (to find other potential cell)
                    if j < area_to_scan[1][1]:  # assure will remain in range of test area for next test (ie, exists cell @ (i, j+1))
                        check_for_index = (i, j + 1)
                        check_for = worksheet.index(check_for_index[0], check_for_index[1])
                        if __check_value(check_for):
                            value_found = check_for
                            index_of_value_found = check_for_index
                    # continue with DOWN test if NOT found somethinng at RIGHT and if this strategy is allowed to be used
                    if down_search_try and not (value_found and index_of_value_found):  # try DOWN test only if method is allowed AND a relevant value not found or value is located in a wrong area
                        # NOTE-LOGIC: test for DOWN cell @(i+1,j) if cell exists in ws range AND if has a value then continue loop (to find other potential cell)
                        if i < area_to_scan[1][0] :  # if still in range if test (ie, exists cell @ (i+1, j))
                            check_for_index = (i + 1, j)
                            check_for = worksheet.index(check_for_index[0], check_for_index[1])
                            if __check_value(check_for):
                                value_found = check_for
                                index_of_value_found = check_for_index
                    # continue with IN-LABEL test if NOT found somethinng at DOWN
                    if not (value_found and index_of_value_found):  # try IN-LABEL test only if a relevant value not found or value is located in a wrong area
                        # NOTE-LOGIC: test for IN-LABEL cell (label is supposed only first word separated by space)
                        check_for = label_value.strip()
                        # keep all except first word (supposed to be label)
                        if len(check_for) > 1:
                            check_for = ' '.join(check_for.split()[1:])  # clean and keep all string except first word
                        else:
                            chech_for = ""
                        if __check_value(check_for):
                            value_found = check_for
                            index_of_value_found = label_location
                # create return dictionary for: label, value found, location of value foun
                ret_val["label_value"] = label_value  # rule: if found a matching label, the label of info found will appear regardles of effective value found status
                ret_val["label_location"] = label_location  # rule: same as for "label_value" key
                if not (value_found and index_of_value_found):  # nothing found, nor @RIGHT, nor @DOWN, so break the loop and try for other poatential cell
                    return ret_val  # will return the empty structure (as initialized)
                if value_found and index_of_value_found:  # if found something will break all loops
                    # all things are ok here (ref effective value), so return found data an, its location
                    ret_val["location"] = index_of_value_found
                    try:  # converting to requested `targeted_type` (will not test is isinstance(obj, Callable) because if not will raise `except` and convert to str)
                        ret_val["value"] = targeted_type(value_found)
                    except:  # noqa: E722
                        ret_val["value"] = str(value_found)
                    return ret_val
    return ret_val  # if get here then will return the empty structure (as initialized)








# #NOTE - ready, test PASS @ 231126 by [piu]
def mk_kv_invoice_items_area(invoice_items_area_xl_format):
    """transform `invoice_items_area` in "canonical JSON format" (as kv pairs).

    Args:
        `invoice_items_area_xl_format`: invoice items area in Excel format (ie, DataFrame with row, col, data).

    Return:
        `invoice_items_area_xl_format`: dictionary with invoice items in Excel format (ie, rows, columns).

    Notes:
        * for ROefact XML model (& plan) see `invoice_files/__model_test_factura_generat_anaf.xml`.
    """
    _invoice_items_data_key = copy.deepcopy(invoice_items_area_xl_format["data"])
    _invoice_items_cols_key = copy.deepcopy(invoice_items_area_xl_format["keycols"])
    _invoice_items_rows_key = copy.deepcopy(invoice_items_area_xl_format["keyrows"])

    _invoice_items_area_json_format = list()
    for _i, _line in enumerate (_invoice_items_rows_key):
        """ identify usual invoice columns: item desc, item UOM, item VAT_percent, item unit_price, item line_value, itemline_ VAT_value
        """
        if True:  # ---- find item quantity column ==> (`cbc_InvoicedQuantity`)
            _col_index = find_str_in_list(["qty", "cant", "quantity"], _invoice_items_cols_key)
            if _col_index is None:  # did not find a suitable column to represent number, so return None probably raising an error
                print(f"[red]***FATAL ERROR - module 'RDINV', function `mk_kv_invoice_items_area(...)`. Cannot find a 'QUANTITY' column in items table. Processing terminated[/]")
                return False
            else:
                _item_quantity = _invoice_items_data_key[_i][_col_index] if isnumber(str(_invoice_items_data_key[_i][_col_index])) else None

        """ #NOTE-1: from here, the following columns are considered "valid" only if a quantity is specified (otherwise is considered an extended description)
        """

        if True:  # ---- find item VAT percent (if exists..., some invoices do noy specify percent itself but just value) ==> (`cbc_Percent`)
            _col_index = find_str_in_list(["cota", "vat %", "% vat"], _invoice_items_cols_key)
            if _col_index is None: # did not find a suitable column to represent number, so return None probably raising an error
                _vat_percent = DEFAULT_VAT_PERCENT if _item_quantity else None  # see #NOTE-1
            else:
                #TODO check_what_is_here_and_if_actual [@231202]:  `_vat_percent` calculation should also consider a simplified invoice where only VAT value is specificed AND THEN SHOULD BE CALCULATED AS_IS in document (see "acciza line on REN... invoice")
                _vat_percent = _invoice_items_data_key[_i][_col_index] if (_invoice_items_data_key[_i][_col_index] is not None) else (DEFAULT_VAT_PERCENT if _item_quantity else None)  # see #TODO-1 check it considering previous comment
                _vat_percent = None if (str(_vat_percent).split() == "") else (float(_vat_percent) if isnumber(str(_vat_percent)) else None)  # finally make it None if remained empty string

        if True:  # ---- find item description / name ==> (`cbc_Name`)
            _col_index = find_str_in_list(["denumire", "name", "nume", "item", "desc"], _invoice_items_cols_key)
            if _col_index is None: # did not find a suitable column to represent number, so return None probably raising an error
                _name_description = DEFAULT_UNKNOWN_ITEM_NAME if _item_quantity else None  # see #NOTE-1
            else:
                _name_description = _invoice_items_data_key[_i][_col_index] if (_invoice_items_data_key[_i][_col_index] is not None) else (DEFAULT_UNKNOWN_ITEM_NAME if _item_quantity else None)  # see #NOTE-1

        if True:  # --- find unit of measure ==> (`cbc_unitCode`)
            _col_index = find_str_in_list(["uom", "um", "masura", "measure"], _invoice_items_cols_key)
            if _col_index is None: # did not find a suitable column to represent number, so return None probably raising an error
                _unif_of_measure = DEFAULT_UNKNOWN_UOM if _item_quantity else None  # see #NOTE-1
            else:
                _unif_of_measure = _invoice_items_data_key[_i][_col_index] if (_invoice_items_data_key[_i][_col_index] is not None) else (DEFAULT_UNKNOWN_UOM if _item_quantity else None)  # see #NOTE-1

        if True:  # --- find unit price ==> (`cbc_PriceAmount`)
            _col_index = find_str_in_list(["price", "pret"], _invoice_items_cols_key)
            if _col_index is None:  # did not find a suitable column to represent number, so return None probably raising an error
                _unit_price = 0 if _item_quantity else None  # see #NOTE-1
            else:
                _tmp = float(_invoice_items_data_key[_i][_col_index]) if isnumber(str(_invoice_items_data_key[_i][_col_index])) else None  # convert it to numer if possible
                _unit_price = _tmp if (_tmp is not None) else (_tmp if _item_quantity else None)  # see #NOTE-1

        if True:  # --- find CURRENCY ==> (`cbc_currencyID`)
            pass # NOTE this will be identifyed in `invoice_header_area` ==> should be changed accordingly

        if True:  # --- calculate line totals ==> (`cbc_LineExtensionAmount`)
            _item_total = None
            if (_item_quantity is not None) and (_unit_price is not None):
                _item_total = round(_item_quantity * _unit_price, 2)
                # line VAT calculation is subject of VAT existence and right calculation as number, otherwise it is set to NULL
                if _item_total and _vat_percent:
                    _item_VAT = float(_item_quantity * _unit_price * _vat_percent)  # this line does not round info to be able to round only the total
                else:
                    _item_VAT = 0.0

        # build dictionary with usual invoice columns (respecting as possible the XSD schemes listed in [`meta_info`][`invoice_XML_schemes`] key)
        _line_info = {
            "cac_InvoiceLine": {
                "cbc_ID": str(_line),
                "cbc_InvoicedQuantity": _item_quantity,
                "cbc_unitCode": None if (_item_quantity is None or str(_item_quantity).split() == "") else _unif_of_measure,
                "cac_Item": {  #-NOTE these are the item specifications (uom, vat)
                    "cbc_Name": str(_name_description),
                    "cac_ClassifiedTaxCategory": {
                        "cbc_Percent": _vat_percent,
                        "cac_TaxScheme": {
                            "cbc_ID": None if (_vat_percent is None or str(_vat_percent).split() == "") else "VAT"
                        }
                    }
                },
                "cac_Price": {
                    "cbc_PriceAmount" : _unit_price,
                    "cbc_currencyID": None if (_item_quantity is None or str(_item_quantity).split() == "") else DEFAULT_CURRENCY
                },
                "cbc_LineExtensionAmount": _item_total,
                "LineVatAmount": None if _item_total is None else _item_VAT,  # line VAT calculation is subject of some Amount existence
            }
        }
        _invoice_items_area_json_format.append(_line_info["cac_InvoiceLine"])
    return copy.deepcopy(_invoice_items_area_json_format)






# #NOTE - ready, test PASS @ 231121 by [piu]
def get_invoice_items_area(worksheet, invoice_items_area_marker, wks_name):
    """get invoice for `invoice_items_area`, process it and return its Excel format.

    Process steps & notes:
        * find invoice items subtable.
        * clean invoice items subtable.
        * extract relevenat data.
        * NOTE: all Excel cell addresses are in `(row, col)` format (ie, Not Excel format like "A:26, C:42, ...")

    Args:
        `worksheet`: the worksheet containing invoice (as object of `pyxllight` library).
        `invoice_items_area_marker`: string with exact marker of invoice items table.
            NOTE: this is the UPPER-LEFT corner and is determined before calling this procedure.
        `wks_name`: the wroksheet name (string) of the `worksheet` object.

    Return:
        `invoice_items_area`: dictionary with invoice items in Excel format (ie, rows, columns).
    """
    # obtain table with invoice items ==> `invoice_items_area`
    invoice_items_area = worksheet.ssd(keycols = invoice_items_area_marker, keyrows = invoice_items_area_marker)
    if (invoice_items_area is None or ((isinstance(invoice_items_area, list)) and len(invoice_items_area) < 1)):  # there was not detected any area candidate to "invoice items / lines", so will exit rasing error
        print(f"[red]***FATAL ERROR - Cannot find any candidate to for invoice ITEMS. Worksheet - \"{wks_name}\" in Module [red] RDINV (code-name: `rdinv`). File processing terminated[/]")
        return False

    #TODO test if list has more items (ie, that means more item tables that will need to be consolidated)
    if isinstance(invoice_items_area, list) and len(invoice_items_area) > 0:
        # NOTE `invoice_items_area` dictionary with keys: "keyrows", "keycols" and "data" (self explanatory)
        invoice_items_area = invoice_items_area[0]  # will suppose found just one AND retain only first one (index [0]) - SEE AFTER TEST with RENware invoice...

    """ CLEANING & CLEARING section
    """
    if True:  # preserve actual rows index in a separated structure (`invoice_items_area["keyrows_index"]`)
        invoice_items_area["keyrows_index"] = list()
        for _tmp_row_index, _tmp_row in enumerate(invoice_items_area["keyrows"]):  # scan all rows and those with empty name/title are first candidates
            invoice_items_area["keyrows_index"].append(_tmp_row_index)
            # clean full empty rows
            if _tmp_row == SYS_FILLED_EMPTY_CELL:
                # inspect all row cells to see if all are empty
                _tmp_test_row_if_full_zero = sum([0 if _i == SYS_FILLED_EMPTY_CELL else 1 for _i in invoice_items_area["data"][_tmp_row_index]])
                if _tmp_test_row_if_full_zero == 0:  # efectivelly delete in subject objects
                    del invoice_items_area["keyrows"][_tmp_row_index]  # drop that row from "keyrows" keyword list
                    # del invoice_items_area["keyrows_index"][_tmp_row_index]  # drop that row from "keyrows" keyword list #NOTE there is no need, you just enumerated this index and dropeed it in the same for-loop step
                    del invoice_items_area["data"][_tmp_row_index]  # drop that row from "data" keyword list
        del invoice_items_area["keyrows_index"]  # cleanup after do job ... :)

    if True:  # clean empty columns: columns without a name will be completly dropped as they are unusable anyway (ie, do not know what to do with them...)
        _tmp_cells_to_drop_in_data_key = list()
        _tmp_items_to_drop_in_keycols_key = list()
        for _tmp_col_index, _tmp_col in enumerate(invoice_items_area["keycols"]):  # scan all cols and those with empty name/title are first candidates
            if _tmp_col == SYS_FILLED_EMPTY_CELL:
                # inspect all col cells to see if all are empty & efectivelly delete in subject objects
                _tmp_items_to_drop_in_keycols_key.append(_tmp_col_index)
                for _data_row_index, _data_row in enumerate(invoice_items_area["data"]):  # scan all rows to find out cells that are part of in subject columns
                    # find out DATA all cells that corresponding to columns to be dropped ==> `_tmp_cells_to_drop_in_data_key: list[(row, col)]`
                    _tmp_tmp = (_data_row_index, _tmp_col_index)
                    _tmp_cells_to_drop_in_data_key.append(_tmp_tmp)
        # drop collected objects (column heads from KEYCOLS & correcsponding cell from DATA)
        for _obj_to_delete in reversed(_tmp_items_to_drop_in_keycols_key):  # from KEYCOLS... (start with last item to not remain "in air" due to deletions :))
            del invoice_items_area["keycols"][_obj_to_delete]
        for _object_to_delete in reversed(_tmp_cells_to_drop_in_data_key):  # from DATA... (start with last item to not remain "in air" due to deletions :))
            del invoice_items_area["data"][_object_to_delete[0]][_object_to_delete[1]]  # drop that col from "data" keyword list

    if True:  # unknown header rows: set as a descriptive using format: `<current line number>.NOTE-<seq>`, where `seq` is an ordered sequence of letters (ie, resulting something like: `1.a, 1.b, ...`)
        _prev_row_number = None
        __letter_seq_idx = 0  # used for next letter sequence (will reset after "a normal one" row (ie, has a value))
        for _crt_row_idx, _crt_row_val in enumerate(invoice_items_area["keyrows"]):
            #
            # if current row is "a normal one", then preserve index and reset letter sequencer
            if _crt_row_val != SYS_FILLED_EMPTY_CELL:
                _prev_row_number = _crt_row_val  # preserve row business index (ie, value shown invoice data not array index)
                __letter_seq_idx = 0   # reset letter sequence generator
            #
            # if current row is "a unknown one" set its header as `<prev line number>.NOTE-<seq>`, where `seq` is an ordered sequence of letters (ie, resulting something like: `1.a, 1.b, ...`)
            if _crt_row_val == SYS_FILLED_EMPTY_CELL:
                _crt_row_seq = f"{_prev_row_number}.NOTE-{ascii_lowercase[__letter_seq_idx]}"  # build a sequence-text for current item (see #NOTE_format)
                # update line header
                invoice_items_area["keyrows"][_crt_row_idx] = str(_crt_row_seq)
                __letter_seq_idx += 1

    if True:  # set back to empty cells that remained to `SYS_FILLED_EMPTY_CELL` (only in `invoice_items_area`)
        invoice_items_area["keycols"] = ["" if _i == SYS_FILLED_EMPTY_CELL else _i for _i in invoice_items_area["keycols"]]  # loop for 'keycols' keyword
        invoice_items_area["keyrows"] = ["" if _i == SYS_FILLED_EMPTY_CELL else _i for _i in invoice_items_area["keyrows"]]  # loop for 'keyrows' keyword
        for _tmp_row_index, _tmp_row in enumerate(invoice_items_area["data"]):  # # loop for 'data' keyword (first loop table rows, "data" key is matrix of lines & cells, ie as [][])
            _tmp_row = ["" if _i == SYS_FILLED_EMPTY_CELL else _i for _i in _tmp_row]
            invoice_items_area["data"][_tmp_row_index] = _tmp_row
    return copy.deepcopy(invoice_items_area)






# #NOTE - ready, test PASS @ 231111 by [piu]
def _get_merged_cells_tobe_changed(file_to_scan, invoice_worksheet_name, keep_cells_of_items_ssd_marker = None):
    """scan Excel file to detect all merged ranges.

    Args:
        `file_to_scan`: the excel file to be scanned.
        `invoice_worksheet_name`: the worksheet to be scanned.
        `keep_cells_of_items_ssd_marker`: tuple with cells that will be marked IN ANY CASE to be preserved:
            * use case: to keep all potential invoice items ssd rows.
            * format: `tuple(row, col, val)` where row & col are relevant here
            * default: `None`

    Return:
        `cells_to_be_changed`: list with cells that need to be chaged in format `(row,col)`.

    Notes:
        * function is intended to be used ONLY internal in this module.
        * use `openpyxl` library to do its job.
    """
    all_detected_ranges = []
    # open Excel file & worksheet
    workbook_opnxl= opnxl.load_workbook(file_to_scan)
    worksheet_opnxl = workbook_opnxl[invoice_worksheet_name]
    all_detected_ranges = worksheet_opnxl.merged_cells.ranges  # get all merged ranges
    _cells_to_be_changed = list()  # will retaing cells that should be marked with SYS_FILLED_EMPTY_CELL
    for _crt_range in all_detected_ranges:
        # range coordinates
        _crt_range_START_COL = _crt_range.bounds[0]
        _crt_range_END_COL = _crt_range.bounds[2]
        _crt_range_START_ROW = _crt_range.bounds[1]
        _crt_range_END_ROW = _crt_range.bounds[3]
        # traverse merged range for all cells in / set flags for first entry and when to completly break loop
        _first_entry = True;
        _full_break = False
        for c in range(_crt_range_START_COL, _crt_range_END_COL + 1):  # traverse all COLS ...
            for r in range(_crt_range_START_ROW, _crt_range_END_ROW + 1):  # traverse all ROWS ...
                _crt_cell_value = worksheet_opnxl.cell(r, c).value
                #print(f"\t/***** processing cell (row,col = {r},{c}) has value {_crt_cell_value}")  #NOTE for debug purposes
                if _first_entry:  # at first pass, see if relevant (ie, not empty or empty string) AND if NOT then break all loops
                    if _crt_cell_value is None:
                        _full_break = True
                        break
                    if isinstance(_crt_cell_value, str) and _crt_cell_value.strip() == "":  # if is a string test if cell is a real empty string
                        _full_break = True
                        break
                    #print(f"\t/***** RELEVANT cell (row,col = {r},{c}) has value {_crt_cell_value}")  #NOTE for debug purposes
                    """ section applicable only when `keep_cells_of_items_ssd_marker` is note None
                        * (r1) in this case, if found a relevant entry in a merged range and and first entry in merged range scan
                        * (r2) to preserve cell @(`row=current_row`, `col=keep_cells_of_items_ssd_marker[1]`)
                        * (r3) if that cell is empty, them add in `_cells_to_be_changed` list to be marked with SYS_FILLED_EMPTY_CELL
                    """
                    if keep_cells_of_items_ssd_marker:
                        _potential_ROW_INDEX_address = (r, keep_cells_of_items_ssd_marker[1])
                        __this_cell_value = worksheet_opnxl.cell(*_potential_ROW_INDEX_address).value
                        if not (__this_cell_value) and (_potential_ROW_INDEX_address not in _cells_to_be_changed):  # apply (r3) rule - see long comment before this section
                            _cells_to_be_changed.append(_potential_ROW_INDEX_address)
                if not _first_entry:  # here the cell has a relevant value, store all next to be marked with SYS_FILLED_EMPTY_CELL
                    _cells_to_be_changed.append((r, c))
                    #print(f"\t/***** cell {(r, c)} marked for '___sys_filled_empty_cell' / [yellow]all list is {_cells_to_be_changed}[/]") #NOTE for debug purposes
                _first_entry = False
            if _full_break:
                break
    return tuple(copy.deepcopy(_cells_to_be_changed))  # always return a tuple as being immutable






# #NOTE - ready, test PASS @ 231127 by [piu]
def _build_meta_info_key(excel_file_to_process: str,
                         invoice_worksheet_name: str,
                         ws_size: list,
                         keyword_for_items_table_marker: str,
                         found_cell: list) -> dict:
    """build meta_info key to preserve processed Excel file meta information: start address, size.

    Notes:
        1: all cell addresses are in format (row, col) and are absolute (ie, valid for whole Excel file) #TODO subject of documentation update.
        2: this function is designed to be used internally by current module (using outside it is not guaranteed for information 'quality').

    Args:
        `excel_file_to_process`: name of file to process as would appear in `meta_info` key.
        `invoice_worksheet_name`: the worksheet name as would appear in `meta_info` key.
        `ws_size`: worksheet size as would appear in `meta_info` key (index 0 max rows, index 1 max columns).
        `keyword_for_items_table_marker`: the content of cell used as start of invoice items subtable as would appear in `meta_info`.
        `found_cell`: position of cell used as start of invoice items subtable as would appear in `meta_info` key (index 0 row, index 1 column).

    Return:
        `meta_info` dictionary built with meta information to be incorpoarted in final invoice dict
    """
    _tmp_meta_info = dict()

    _tmp_meta_info["file"] = os.path.basename(excel_file_to_process)
    _tmp_meta_info["file_CRC"] = "...file CRC (uniquely identify the invoice file used)"  #TODO to be done... #NOTE this calculation should be done as last step after final XLSX file writing
    _tmp_meta_info["last_processing_time"] = datetime.now(timezone.utc).isoformat()  # set to ISO 8601 format
    _tmp_meta_info["invoice_worksheet"] = invoice_worksheet_name
    _tmp_meta_info["invoice_max_rows"] = ws_size[0]
    _tmp_meta_info["invoice_max_cols"] = ws_size[1]
    _tmp_meta_info["items_table_start_marker"] = keyword_for_items_table_marker
    _tmp_meta_info["items_table_start_cell"] = (found_cell[0], found_cell[1])
    _tmp_meta_info["invoice_XML_schemes"] = {
        "xmlns": "urn:oasis:names:specification:ubl:schema:xsd:Invoice-2",
        "xmlns:cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        "xmlns:cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
        "xmlns:ns4": "urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2",
        "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
        "xsi:schemaLocation": "urn:oasis:names:specification:ubl:schema:xsd:Invoice-2 http://docs.oasis-open.org/ubl/os-UBL-2.1/xsd/maindoc/UBL-Invoice-2.1.xsd"
    }
    _tmp_meta_info["map_JSONkeys_XMLtags"] = [  # list of tuple(JSONkey: str, XMLtag: str) #TODO subject of documentation update
        ("cac_InvoiceLine", "cac:InvoiceLine"),
        ("cac_Item", "cac:Item"),
        ("cac_ClassifiedTaxCategory", "cac:ClassifiedTaxCategory"),
        ("cac_TaxScheme", "cac:TaxScheme"),
        ("cac_Price", "cac:Price"),
        ("cbc_ID", "cbc:ID"),
        ("cbc_InvoicedQuantity", "cbc:InvoicedQuantity"),
        ("cbc_unitCode", "cbc:unitCode"),  # this is attribute of tag InvoicedQuantity
        ("cbc_Name", "cbc:Name"),
        ("cbc_Percent", "cbc:Percent"),
        ("cbc_PriceAmount", "cbc:PriceAmount"),
        ("cbc_currencyID", "cbc:currencyID"),  # this is attribute of more tags (all monetary tags)
        ("cbc_LineExtensionAmount", "cbc:LineExtensionAmount"),
        ("cbc_ID", "cbc:ID"),  # invoice number
        ("cbc_DocumentCurrencyCode", "cbc:DocumentCurrencyCode"),  # invoice currency
        ("cbc_IssueDate", "cbc:IssueDate"),  # invoice issue date
        ("cac_AccountingCustomerParty", "cac:AccountingCustomerParty"),  # invoice customer information - MASTER RECORD
        ("cac_Party", "cac:Party"),  # invoice customer details ref Parner info (legal, address, ...) - DETAIL L1 RECORD
        ("cac_PartyLegalEntity", "cac:PartyLegalEntity"),  # invoice customer inforation - DETAIL L2 RECORD
        ("cbc_CompanyID", "cbc:CompanyID"),  # invoice customer inforation - DETAIL L3 RECORD
        ("cbc_RegistrationName", "cbc:RegistrationName"),  # invoice customer inforation - DETAIL L3 RECORD
        ("cac_PostalAddress", "cac:PostalAddress"),  # invoice customer postal address info - DETAIL L2 RECORD
        ("cbc_StreetName", "cbc:StreetName"),  # invoice customer inforation - DETAIL L3 RECORD
        ("cbc_CityName", "cbc:CityName"),  # invoice customer inforation - DETAIL L3 RECORD
        ("cbc_PostalZone", "cbc:PostalZone"),  # invoice customer inforation - DETAIL L3 RECORD
        ("cac_Country", "cac:Country"),  # invoice customer inforation - DETAIL L3 RECORD
        ("cbc_IdentificationCode", "cbc:IdentificationCode"),  # invoice customer inforation - DETAIL L3 RECORD
        ("LineVatAmount", None),  # line / item total VAT. Has no correspondent in XML schema - DETAIL L3 RECORD
        ("cac_Contact", "cac:Contact"),  #  customer contact information: email, phone - DETAIL L2 RECORD
        ("cbc_Telephone", "cbc:Telephone"),  # customer phone -- DETAIL L3 RECORD
        ("cbc_ElectronicMail", "cbc:ElectronicMail"),  # customer email - DETAIL L3 RECORD
        ("RegCom", None),  # customer commerce register number (company legal registration number). Has no correspondent in XML schema - DETAIL L3 RECORD
        ("Bank", None),  # customer bank. Has no correspondent in XML schema - DETAIL L3 RECORD
        ("IBAN", None),  # customer bank account number (IBAN). Has no correspondent in XML schema - DETAIL L3 RECORD
        #TODO ...here to add items ref `cac_PostalAddress` - DETAIL L3 RECORDS
    ]

    return copy.deepcopy(_tmp_meta_info)




