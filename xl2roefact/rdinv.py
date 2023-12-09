#!./.wenv_xl2roefact/bin/python3
"""  RDINV - modul de procesare a fisierului format XLSX ce contine factura si colectare a datelor aferente

    Identification:
        code-name: `rdinv`
        copyright: (c) 2023 RENWare Software Systems
        author: Petre Iordanescu (petre.iordanescu@gmail.com)

    Specifications:
        document: `110-SRE-api_to_roefact_requirements.md` section `Componenta xl2roefact`
        INTRARI: fisier format XLSX ce contine factura emisa (cod: `f-XLSX`)
        IESIRI: fisier format JSON imagine a datelor facturii (cod: `f-JSON`)
"""

import os, sys
from datetime import datetime, timezone, tzinfo
from rich import print
import copy
from rich.pretty import pprint
from string import ascii_lowercase
import json
from libutils import isnumber, find_str_in_list  # application misc/general utilities
import config_settings  # application configuration parameters
import pylightxl as xl
import openpyxl as opnxl


SYS_FILLED_EMPTY_CELL = "_sys_keep_cell"  # this is not a changeale constant


def rdinv(file_to_process: str, invoice_worksheet_name: str = None, debug_info: bool = False):
    """ main function of RDINV module

    Arguments:
        - `file_to_process`: the invoice file (exact file with path)
        - `invoice_worksheet_name`: the worksheet containing invoice
        - `debug_info`: bool to show debugging information, default `False`

    Return:
        - `invoice`: the invoice extracted information from Excel file as
                     `dict(meta_info: dict, invoice_header_area: dict, invoice_items_area: dict, invoice_footer_area: dict)`  #TODO subject of documentation update

    Important variables:
        - `db`: htxl object with invoice EXCEL (as a whole)
        - `ws`: pylightxl object with invoice WORKSHEET
    """

    print(f"*** Module [red]rdinv[/] started at {datetime.now()} to process file [green]{os.path.split(file_to_process)[1]}[/] (full path: {file_to_process})")

    # read Excel file with Invoice data
    try:
        db = xl.readxl(fn=file_to_process)
    except:
        print(f"[red]---***FATAL ERROR - Cannot open Excel file {file_to_process} (possible problems: file corrupted, wrong type only XLSX accetpted, file does not exists or was deleted, operating system vilotation) in Module [red] RDINV (code-name: `rdinv`). File processing terminated[/]")
        return False

    # read the workshet with Invoice data
    if invoice_worksheet_name is None:  # if parameter `invoice_worksheet_name` not specified try to open first worksheet from Excel worksheets - order is given by worksheets order in Excel file
        list_of_excel_worksheets = db.ws_names
        print(f"[yellow]INFO note:[/] `rdinv` module, no worksheet specified so will open [cyan]'{list_of_excel_worksheets[0]}'[/]")
        invoice_worksheet_name = list_of_excel_worksheets[0]

    try:
        ws = db.ws(invoice_worksheet_name)
    except:
        print(f"[red]***FATAL ERROR - Cannot open Excel specified Worksheet \"{invoice_worksheet_name}\" in Module [red] RDINV (code-name: `rdinv`). File processing terminated[/]")
        return False


    """ #NOTE: section for search of `invoice_items_area` (ie `pylightxl.ssd` object)
        - how: search the cell containg text fragments --> get text from that cell --> use it as `ssd()` parameter
        - result: `keyword_for_items_table_marker` = string marker to search for in oredr to isolate `invoice_items_area`
        - NOTE: partial result: `_found_cell = (row, col, val)`
    """
    _ws_max_rows, _ws_max_cols = ws.size[0], ws.size[1]
    _FOUND_RELEVANT_CELL = False
    for _crt_row in range(1, _ws_max_rows + 1):  # traverse all rows (start from 1 as Excel style)
        for _crt_col in range(1, _ws_max_cols + 1):  # traverse all cols (start from 1 as Excel style)
            _crt_cell_val = ws.index(_crt_row, _crt_col)
            if (_crt_cell_val == "") or (_crt_cell_val == SYS_FILLED_EMPTY_CELL) or (_crt_cell_val is None):  # skip empty cells and continue with next cells
                continue
            # search for all strings from config_settings.INVOICE_ITEMS_SUBTABLE_MARKER
            _cell_val_to_test = str(_crt_cell_val).lower()
            for i in config_settings.INVOICE_ITEMS_SUBTABLE_MARKER:   # search in current cell contains one the strings potential to identify items subtable
                if i in _cell_val_to_test:
                    _found_cell = (_crt_row, _crt_col, _crt_cell_val)
                    _FOUND_RELEVANT_CELL = True
                    break  # found a relevant cell ==> exit
                else:
                    continue
            if _FOUND_RELEVANT_CELL:
                break
        if _FOUND_RELEVANT_CELL:
            break
    if not _FOUND_RELEVANT_CELL:
        print(f"[red]***FATAL ERROR - Cannot find a relevant cell where invoice items table start (basically containing string \" crt\"). File processing terminated[/]")
        return False
    keyword_for_items_table_marker = _found_cell[2]  #NOTE here you have `_found_cell = (row, col, val)` so can set variable `keyword_for_items_table_marker`

    # detect all cells that should be changed to SYS_FILLED_EMPTY_CELL (these are cells id merged groups where first cell in merged group is relevant (diff from empty))
    detected_cells_which_will_be_fake_filled = _get_merged_cells_tobe_changed(
        file_to_scan=file_to_process,
        invoice_worksheet_name=invoice_worksheet_name,
        keep_cells_of_items_ssd_marker=_found_cell)  # call specify that cells with some description but not real invoice lines but probably just details / supplimentary explanations to last real invoce line, will be kept (to be preserved in final invoice as: extended description)
    # scan all detectected cell and change them
    for _cell_index in detected_cells_which_will_be_fake_filled:
        _cell_row = _cell_index[0]
        _cell_col = _cell_index[1]
        ws.update_index(row = _cell_row, col = _cell_col, val = SYS_FILLED_EMPTY_CELL)


    """ #NOTE: section for solve `invoice_items_area` in 2 steps:
        - first process it as Excel format (row & colomns datale (aka Data Frame))
        - transform it in "canonical JSON format" (as kv pairs)
    """
    # process invoice to detect its items / lines ('invoice_items_area'), clean and extract data
    invoice_items_area = _get_invoice_items_area(
        worksheet=ws,
        invoice_items_area_marker=keyword_for_items_table_marker,
        wks_name=invoice_worksheet_name
    )
    # transform `invoice_items_area` in "canonical JSON format" (as kv pairs)
    invoice_items_as_kv_pairs = __mk_kv_invoice_items_area(invoice_items_area_xl_format=invoice_items_area)


    """ #NOTE: section for solve `invoice_items_area`
        - #TODO ...hereuare... ...short plan here... owner, partner, invoice identification, currency, invoice id/number, issued date
        - @final write them to:
            - (1) ==> `invoice["excel_original_data"]["invoice_header_area"]` then
            - (2) transform them to canonocal form (as @invoice lines, see line ~122) ==> `invoice["Invoice"]`
    """
    #TODO ...hereuare...


    # preserve processed Excel file meta information: start address, size.
    meta_info = _build_meta_info_key(
        excel_file_to_process=file_to_process,
        invoice_worksheet_name=invoice_worksheet_name,
        ws_size=tuple(ws.size),
        keyword_for_items_table_marker=keyword_for_items_table_marker,
        found_cell=tuple(_found_cell))

    # build final structure to be returned (`invoice`) - MAIN OBJECTIVE of this function
    invoice = {
        "Invoice": {
            "cac_InvoiceLine": [_i for _i in invoice_items_as_kv_pairs]  # `invoice_items_as_kv_pairs` is a list of dicts with keys as XML/XSD RO E-Fact standard
        },
        "meta_info": copy.deepcopy(meta_info),
        "excel_original_data": dict(
            invoice_items_area = copy.deepcopy(invoice_items_area),
            invoice_header_area = "...to be done",  #TODO to be done...
            invoice_footer_area = "...to be done"  #TODO to be done...
        )
    }

    # write `invoice` dict to `f-JSON`
    """ useful NOTE(s):
        - ref `f-JSON` file, see doc: `https://apitoroefact.renware.eu/commercial_agreement/110-SRE-api_to_roefact_requirements.html#vedere-de-ansamblu-a-solutiei`
        - create `f-JSON` filename as Excel filename but with json extention
        - helpers:
                - os.path.split gets @[0] directory & @[1] filename
                - os.path.splitext @[0] filename w/o ext, @[1] extention woth dot char included
    """
    _fjson_filename = os.path.splitext(os.path.basename(file_to_process))[0] + ".json"
    _fjson_fileobject = os.path.join(os.path.split(file_to_process)[0], _fjson_filename)
    with open(_fjson_fileobject, 'w', encoding='utf-8') as _f:
        json.dump(invoice, _f, ensure_ascii = False, indent = 4)
    print(f"[yellow]INFO note:[/] `rdinv` module, written invoice JSON data to: [green]{_fjson_fileobject}[/]")


    #TODO check for more TODOs, clean &&-->
    #TODO wip...(@231125) TRANSFORM JSON FILE from Excel (row,col) format in a relational one (but respecting ROefact tags from used scheme)

    '''  #NOTE [@231208] moved in `xl2roefact` as debog-verbose option  #FIXME drop me after a while...
    if debug_info:  # NOTE DEBUG area print
        print(f"[yellow]DEBUG note:[/] content of `invoice` data dictionary:")
        pprint(invoice)
        print()
    '''

    return copy.deepcopy(invoice)






# #NOTE - ready, test PASS @ 231126 by [piu]
def __mk_kv_invoice_items_area(invoice_items_area_xl_format):
    """ transform `invoice_items_area` in "canonical JSON format" (as kv pairs)

    Arguments:
        - `invoice_items_area_xl_format` - invoice items area in Excel format (ie, DataFrame with row, col, data)

    Return:
        - `invoice_items_area_xl_format` - dictionary with invoice items in Excel format (ie, rows, columns)
    Notes:
        - for ROefact XML model (& plan) see `invoice_files/__model_test_factura_generat_anaf.xml`
    """
    _invoice_items_data_key = copy.deepcopy(invoice_items_area_xl_format["data"])
    _invoice_items_cols_key = copy.deepcopy(invoice_items_area_xl_format["keycols"])
    _invoice_items_rows_key = copy.deepcopy(invoice_items_area_xl_format["keyrows"])

    _invoice_items_area_json_format = list()
    for _i, _line in enumerate (_invoice_items_rows_key):
        """ identify usual invoice columns: item desc, item UOM, item VAT_percent, item unit_price, item line_value, itemline_ VAT_value
        """
        if True:  # ---- find item quantity column ==> (`cbc_InvoicedQuantity`)
            _col_index = find_str_in_list(["qty", "cant", "quantity"], _invoice_items_cols_key)
            if _col_index is None:  # did not find a suitable column to represent number, so return None probably raising an error
                print(f"[red]***FATAL ERROR - module 'RDINV', function `__mk_kv_invoice_items_area(...)`. Cannot find a 'QUANTITY' column in items table. Processing terminated[/]")
                return False
            else:
                _item_quantity = _invoice_items_data_key[_i][_col_index] if isnumber(str(_invoice_items_data_key[_i][_col_index])) else None

        """ #NOTE-1: from here, the following columns are considered "valid" only if a quantity is specified (otherwise is considered an extended description)
        """

        if True:  # ---- find item VAT percent (if exists..., some invoices do noy specify percent itself but just value) ==> (`cbc_Percent`)
            _col_index = find_str_in_list(["cota", "vat %", "% vat"], _invoice_items_cols_key)
            if _col_index is None: # did not find a suitable column to represent number, so return None probably raising an error
                _vat_percent = config_settings.DEFAULT_VAT_PERCENT if _item_quantity else None  # see #NOTE-1
            else:
                #FIXME `_vat_percent` calculation should also consider a simplified invoice where only VAT value is specificed AND THEN SHOULD BE CALCULATED AS_IS in document (see "acciza line on REN... invoice")
                _vat_percent = _invoice_items_data_key[_i][_col_index] if (_invoice_items_data_key[_i][_col_index] is not None) else (config_settings.DEFAULT_VAT_PERCENT if _item_quantity else None)  # see #NOTE-1 #FIXME fix it considering previous comment
                _vat_percent = None if (str(_vat_percent).split() == "") else (float(_vat_percent) if isnumber(str(_vat_percent)) else None)  # finally make it None if remained empty string

        if True:  # ---- find item description / name ==> (`cbc_Name`)
            _col_index = find_str_in_list(["denumire", "name", "nume", "item", "desc"], _invoice_items_cols_key)
            if _col_index is None: # did not find a suitable column to represent number, so return None probably raising an error
                _name_description = config_settings.DEFAULT_UNKNOWN_ITEM_NAME if _item_quantity else None  # see #NOTE-1
            else:
                _name_description = _invoice_items_data_key[_i][_col_index] if (_invoice_items_data_key[_i][_col_index] is not None) else (config_settings.DEFAULT_UNKNOWN_ITEM_NAME if _item_quantity else None)  # see #NOTE-1

        if True:  # --- find unit of measure ==> (`cbc_unitCode`)
            _col_index = find_str_in_list(["uom", "um", "masura", "measure"], _invoice_items_cols_key)
            if _col_index is None: # did not find a suitable column to represent number, so return None probably raising an error
                _unif_of_measure = config_settings.DEFAULT_UNKNOWN_UOM if _item_quantity else None  # see #NOTE-1
            else:
                _unif_of_measure = _invoice_items_data_key[_i][_col_index] if (_invoice_items_data_key[_i][_col_index] is not None) else (config_settings.DEFAULT_UNKNOWN_UOM if _item_quantity else None)  # see #NOTE-1

        if True:  # --- find unit price ==> (`cbc_PriceAmount`)
            _col_index = find_str_in_list(["price", "pret"], _invoice_items_cols_key)
            if _col_index is None:  # did not find a suitable column to represent number, so return None probably raising an error
                _unit_price = 0 if _item_quantity else None  # see #NOTE-1
            else:
                _tmp = float(_invoice_items_data_key[_i][_col_index]) if isnumber(str(_invoice_items_data_key[_i][_col_index])) else None  # convert it to numer if possible
                _unit_price = _tmp if (_tmp is not None) else (_tmp if _item_quantity else None)  # see #NOTE-1

        if True:  # --- find CURRENCY ==> (`cbc_currencyID`) (#FIXME this will be identifyed in `invoice_header_area` ==> should be changed accordingly)
            pass #FIXME this will be identifyed in `invoice_header_area` ==> should be changed accordingly

        if True:  # --- calculate line totals ==> (`cbc_LineExtensionAmount`)
            _item_total = None
            if (_item_quantity is not None) and (_unit_price is not None):
                _item_total = round(_item_quantity * _unit_price, 2)

        # build dictionary with usual invoice columns (respecting as possible the XSD schemes listed in [`meta_info`][`invoice_XML_schemes`] key)
        _line_info = {
            "cac_InvoiceLine": {
                "cbc_ID": str(_line),
                "cbc_InvoicedQuantity": _item_quantity,
                "cbc_unitCode": None if (_item_quantity is None or str(_item_quantity).split() == "") else _unif_of_measure,
                "cac_Item": {  #-NOTE these are the item specifications (uom, vat)
                    "cbc_Name": str(_name_description),
                    "cac_ClassifiedTaxCategory": {
                        "cbc_Percent": _vat_percent,
                        "cac_TaxScheme": {
                            "cbc_ID": None if (_vat_percent is None or str(_vat_percent).split() == "") else "VAT"
                        }
                    }
                },
                "cac_Price": {
                    "cbc_PriceAmount" : _unit_price,
                    "cbc_currencyID": None if (_item_quantity is None or str(_item_quantity).split() == "") else config_settings.DEFAULT_CURRENCY  #FIXME this will be identifyed in `invoice_header_area` ==> should be changed accordingly
                },
                "cbc_LineExtensionAmount": _item_total
            }
        }
        _invoice_items_area_json_format.append(_line_info["cac_InvoiceLine"])
    return copy.deepcopy(_invoice_items_area_json_format)






# #NOTE - ready, test PASS @ 231121 by [piu]
def _get_invoice_items_area(worksheet, invoice_items_area_marker, wks_name):
    """ get invoice for `invoice_items_area`, process it and return its Excel format

    Description:
        - find invoice items subtable
        - clean invoice items subtable
        - extract relevenat data
        - NOTE: all Excel cell addresses are in `(row, col)` format (ie, Not Excel format like "A:26, C:42, ...")

    Arguments:
        - `worksheet` - the worksheet containing invoice (as object of `pyxllight` library)
        - `invoice_items_area_marker` - string with exact marker of invoice items table
            NOTE: this is the UPPER-LEFT corner and is determined before calling this procedure
        - `wks_name` the wroksheet name (string) of the `worksheet` object

    Return:
        - `invoice_items_area` - dictionary with invoice items in Excel format (ie, rows, columns)
    """
    # obtain table with invoice items ==> `invoice_items_area`
    invoice_items_area = worksheet.ssd(keycols = invoice_items_area_marker, keyrows = invoice_items_area_marker)
    if (invoice_items_area is None or ((isinstance(invoice_items_area, list)) and len(invoice_items_area) < 1)):  # there was not detected any area candidate to "invoice items / lines", so will exit rasing error
        print(f"[red]***FATAL ERROR - Cannot find any candidate to for invoice ITEMS. Worksheet - \"{wks_name}\" in Module [red] RDINV (code-name: `rdinv`). File processing terminated[/]")
        return False

    #TODO test if list has more items (ie, that means more item tables that will need to be consolidated)
    if isinstance(invoice_items_area, list) and len(invoice_items_area) > 0:
        # NOTE `invoice_items_area` dictionary with keys: "keyrows", "keycols" and "data" (self explanatory)
        invoice_items_area = invoice_items_area[0]  # will suppose found just one AND retain only first one (index [0]) - SEE AFTER TEST with RENware invoice...

    """ CLEANING & CLEARING section
    """
    if True:  # preserve actual rows index in a separated structure (`invoice_items_area["keyrows_index"]`)
        invoice_items_area["keyrows_index"] = list()
        for _tmp_row_index, _tmp_row in enumerate(invoice_items_area["keyrows"]):  # scan all rows and those with empty name/title are first candidates
            invoice_items_area["keyrows_index"].append(_tmp_row_index)
            # clean full empty rows
            if _tmp_row == SYS_FILLED_EMPTY_CELL:
                # inspect all row cells to see if all are empty
                _tmp_test_row_if_full_zero = sum([0 if _i == SYS_FILLED_EMPTY_CELL else 1 for _i in invoice_items_area["data"][_tmp_row_index]])
                if _tmp_test_row_if_full_zero == 0:  # efectivelly delete in subject objects
                    del invoice_items_area["keyrows"][_tmp_row_index]  # drop that row from "keyrows" keyword list
                    # del invoice_items_area["keyrows_index"][_tmp_row_index]  # drop that row from "keyrows" keyword list #NOTE there is no need, you just enumerated this index and dropeed it in the same for-loop step
                    del invoice_items_area["data"][_tmp_row_index]  # drop that row from "data" keyword list
        del invoice_items_area["keyrows_index"]  # cleanup after do job ... :)

    if True:  # clean empty columns: columns without a name will be completly dropped as they are unusable anyway (ie, do not know what to do with them...)
        _tmp_cells_to_drop_in_data_key = list()
        _tmp_items_to_drop_in_keycols_key = list()
        for _tmp_col_index, _tmp_col in enumerate(invoice_items_area["keycols"]):  # scan all cols and those with empty name/title are first candidates
            if _tmp_col == SYS_FILLED_EMPTY_CELL:
                # inspect all col cells to see if all are empty & efectivelly delete in subject objects
                _tmp_items_to_drop_in_keycols_key.append(_tmp_col_index)
                for _data_row_index, _data_row in enumerate(invoice_items_area["data"]):  # scan all rows to find out cells that are part of in subject columns
                    # find out DATA all cells that corresponding to columns to be dropped ==> `_tmp_cells_to_drop_in_data_key: list[(row, col)]`
                    _tmp_tmp = (_data_row_index, _tmp_col_index)
                    _tmp_cells_to_drop_in_data_key.append(_tmp_tmp)
        # drop collected objects (column heads from KEYCOLS & correcsponding cell from DATA)
        for _obj_to_delete in reversed(_tmp_items_to_drop_in_keycols_key):  # from KEYCOLS... (start with last item to not remain "in air" due to deletions :))
            del invoice_items_area["keycols"][_obj_to_delete]
        for _object_to_delete in reversed(_tmp_cells_to_drop_in_data_key):  # from DATA... (start with last item to not remain "in air" due to deletions :))
            del invoice_items_area["data"][_object_to_delete[0]][_object_to_delete[1]]  # drop that col from "data" keyword list

    if True:  # unknown header rows: set as a descriptive using format: `<current line number>.NOTE-<seq>`, where `seq` is an ordered sequence of letters (ie, resulting something like: `1.a, 1.b, ...`)
        _prev_row_number = None
        __letter_seq_idx = 0  # used for next letter sequence (will reset after "a normal one" row (ie, has a value))
        for _crt_row_idx, _crt_row_val in enumerate(invoice_items_area["keyrows"]):
            #
            # if current row is "a normal one", then preserve index and reset letter sequencer
            if _crt_row_val != SYS_FILLED_EMPTY_CELL:
                _prev_row_number = _crt_row_val  # preserve row business index (ie, value shown invoice data not array index)
                __letter_seq_idx = 0   # reset letter sequence generator
            #
            # if current row is "a unknown one" set its header as `<prev line number>.NOTE-<seq>`, where `seq` is an ordered sequence of letters (ie, resulting something like: `1.a, 1.b, ...`)
            if _crt_row_val == SYS_FILLED_EMPTY_CELL:
                _crt_row_seq = f"{_prev_row_number}.NOTE-{ascii_lowercase[__letter_seq_idx]}"  # build a sequence-text for current item (see #NOTE_format)
                # update line header
                invoice_items_area["keyrows"][_crt_row_idx] = str(_crt_row_seq)
                __letter_seq_idx += 1

    if True:  # set back to empty cells that remained to `SYS_FILLED_EMPTY_CELL` (only in `invoice_items_area`)
        invoice_items_area["keycols"] = ["" if _i == SYS_FILLED_EMPTY_CELL else _i for _i in invoice_items_area["keycols"]]  # loop for 'keycols' keyword
        invoice_items_area["keyrows"] = ["" if _i == SYS_FILLED_EMPTY_CELL else _i for _i in invoice_items_area["keyrows"]]  # loop for 'keyrows' keyword
        for _tmp_row_index, _tmp_row in enumerate(invoice_items_area["data"]):  # # loop for 'data' keyword (first loop table rows, "data" key is matrix of lines & cells, ie as [][])
            _tmp_row = ["" if _i == SYS_FILLED_EMPTY_CELL else _i for _i in _tmp_row]
            invoice_items_area["data"][_tmp_row_index] = _tmp_row
    return copy.deepcopy(invoice_items_area)






# #NOTE - ready, test PASS @ 231111 by [piu]
def _get_merged_cells_tobe_changed(file_to_scan, invoice_worksheet_name, keep_cells_of_items_ssd_marker = None):
    """ scan Excel file to detect all merged ranges

    Arguments:
        - `file_to_scan`: the excel file to be scanned
        - `invoice_worksheet_name`: the worksheet to be scanned
        - `keep_cells_of_items_ssd_marker`: tuple with cells that will be marked IN ANY CASE to be preserved
            - use case: to keep all potential invoice items ssd rows
            - format: `tuple(row, col, val)` where row & col are relevant here
            - default: `None`

    Return:
        - `cells_to_be_changed`: list with cells that need to be chaged in format `(row,col)`

    Notes:
        - function is intended to be used ONLY internal in this module
        - use `openpyxl` library to do its job
    """
    all_detected_ranges = []
    # open Excel file & worksheet
    workbook_opnxl= opnxl.load_workbook(file_to_scan)
    worksheet_opnxl = workbook_opnxl[invoice_worksheet_name]
    all_detected_ranges = worksheet_opnxl.merged_cells.ranges  # get all merged ranges
    _cells_to_be_changed = list()  # will retaing cells that should be marked with SYS_FILLED_EMPTY_CELL
    for _crt_range in all_detected_ranges:
        # range coordinates
        _crt_range_START_COL = _crt_range.bounds[0]
        _crt_range_END_COL = _crt_range.bounds[2]
        _crt_range_START_ROW = _crt_range.bounds[1]
        _crt_range_END_ROW = _crt_range.bounds[3]
        # traverse merged range for all cells in / set flags for first entry and when to completly break loop
        _first_entry = True;
        _full_break = False
        for c in range(_crt_range_START_COL, _crt_range_END_COL + 1):  # traverse all COLS ...
            for r in range(_crt_range_START_ROW, _crt_range_END_ROW + 1):  # traverse all ROWS ...
                _crt_cell_value = worksheet_opnxl.cell(r, c).value
                #print(f"\t/***** processing cell (row,col = {r},{c}) has value {_crt_cell_value}")  #NOTE for debug purposes
                if _first_entry:  # at first pass, see if relevant (ie, not empty or empty string) AND if NOT then break all loops
                    if _crt_cell_value is None:
                        _full_break = True
                        break
                    if isinstance(_crt_cell_value, str) and _crt_cell_value.strip() == "":  # if is a string test if cell is a real empty string
                        _full_break = True
                        break
                    #print(f"\t/***** RELEVANT cell (row,col = {r},{c}) has value {_crt_cell_value}")  #NOTE for debug purposes
                    """ section applicable only when `keep_cells_of_items_ssd_marker` is note None
                        * (r1) in this case, if found a relevant entry in a merged range and and first entry in merged range scan
                        * (r2) to preserve cell @(`row=current_row`, `col=keep_cells_of_items_ssd_marker[1]`)
                        * (r3) if that cell is empty, them add in `_cells_to_be_changed` list to be marked with SYS_FILLED_EMPTY_CELL
                    """
                    if keep_cells_of_items_ssd_marker:
                        _potential_ROW_INDEX_address = (r, keep_cells_of_items_ssd_marker[1])
                        __this_cell_value = worksheet_opnxl.cell(*_potential_ROW_INDEX_address).value
                        if not (__this_cell_value) and (_potential_ROW_INDEX_address not in _cells_to_be_changed):  # apply (r3) rule - see long comment before this section
                            _cells_to_be_changed.append(_potential_ROW_INDEX_address)
                if not _first_entry:  # here the cell has a relevant value, store all next to be marked with SYS_FILLED_EMPTY_CELL
                    _cells_to_be_changed.append((r, c))
                    #print(f"\t/***** cell {(r, c)} marked for '___sys_filled_empty_cell' / [yellow]all list is {_cells_to_be_changed}[/]") #NOTE for debug purposes
                _first_entry = False
            if _full_break:
                break
    return tuple(copy.deepcopy(_cells_to_be_changed))  # always return a tuple as being immutable






# #NOTE - ready, test PASS @ 231127 by [piu]
def _build_meta_info_key(excel_file_to_process: str,
                         invoice_worksheet_name: str,
                         ws_size: list,
                         keyword_for_items_table_marker: str,
                         found_cell: list) -> dict:
    """ build meta_info key to preserve processed Excel file meta information: start address, size.

    NOTE(s):
        - all cell addresses are in format (row, col) and are absolute (ie, valid for whole Excel file) #TODO subject of documentation update
        - this function is designed to be used internally by current module (using outside it is not guaranteed for information 'quality')

    Arguments:
        - `excel_file_to_process`: name of file to process as would appear in `meta_info` key
        - `invoice_worksheet_name`: the worksheet name as would appear in `meta_info` key
        - `ws_size`: worksheet size as would appear in `meta_info` key (index 0 max rows, index 1 max columns)
        - `keyword_for_items_table_marker`: the content of cell used as start of invoice items subtable as would appear in `meta_info`
        - `found_cell`: position of cell used as start of invoice items subtable as would appear in `meta_info` key (index 0 row, index 1 column)

    Return:
        - `meta_info` dictionary built with meta information to be incorpoarted in final invoice dict
    """
    _tmp_meta_info = dict()

    _tmp_meta_info["file"] = os.path.basename(excel_file_to_process)
    _tmp_meta_info["file_CRC"] = "...file CRC (uniquely identify the invoice file used)"  #TODO to be done... #NOTE this calculation should be done as last step after final XLSX file writing
    _tmp_meta_info["last_processing_time"] = datetime.now(timezone.utc).isoformat()  # set to ISO 8601 format
    _tmp_meta_info["invoice_worksheet"] = invoice_worksheet_name
    _tmp_meta_info["invoice_max_rows"] = ws_size[0]
    _tmp_meta_info["invoice_max_cols"] = ws_size[1]
    _tmp_meta_info["items_table_start_marker"] = keyword_for_items_table_marker
    _tmp_meta_info["items_table_start_cell"] = (found_cell[0], found_cell[1])
    _tmp_meta_info["invoice_XML_schemes"] = {
        "xmlns": "urn:oasis:names:specification:ubl:schema:xsd:Invoice-2",
        "xmlns:cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        "xmlns:cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
        "xmlns:ns4": "urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2",
        "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
        "xsi:schemaLocation": "urn:oasis:names:specification:ubl:schema:xsd:Invoice-2 http://docs.oasis-open.org/ubl/os-UBL-2.1/xsd/maindoc/UBL-Invoice-2.1.xsd"
    }
    _tmp_meta_info["map_JSONkeys_XMLtags"] = [  # list of tuple(JSONkey: str, XMLtag: str) #TODO subject of documentation update
        ("cac_InvoiceLine", "cac:InvoiceLine"),
        ("cac_Item", "cac:Item"),
        ("cac_ClassifiedTaxCategory", "cac:ClassifiedTaxCategory"),
        ("cac_TaxScheme", "cac:TaxScheme"),
        ("cac_Price", "cac:Price"),
        ("cbc_ID", "cbc:ID"),
        ("cbc_InvoicedQuantity", "cbc:InvoicedQuantity"),
        ("cbc_unitCode", "cbc:unitCode"),  # this is attribute of tag InvoicedQuantity
        ("cbc_Name", "cbc:Name"),
        ("cbc_Percent", "cbc:Percent"),
        ("cbc_PriceAmount", "cbc:PriceAmount"),
        ("cbc_currencyID", "cbc:currencyID"),  # this is attribute of more tags (all monetary tags)
        ("cbc_LineExtensionAmount", "cbc:LineExtensionAmount")
    ]

    return copy.deepcopy(_tmp_meta_info)




